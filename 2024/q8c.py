import re
import time
from urllib.parse import urlencode

from bs4 import BeautifulSoup
from selenium import webdriver

import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait

from selenium.webdriver.support import expected_conditions as EC

file1 = open("q8_log.txt", "r")
lines = file1.readlines()

songs = []
for line in lines:
    songs.append(line.rstrip())

# read from excel

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("top2000.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

songs = []

skip = False
colums = list(dataframe1.iter_cols(1, dataframe1.max_column))
for row in range(1, dataframe1.max_row):
    if skip:
        if colums[1][row].value == "Boogie Wonderland":
            skip = False

        continue


    song = f"{colums[1][row].value.replace('(Albumversie)', '')} - {colums[2][row].value}"
    songs.append(song)


print(len(songs))


for i, song_title_artist in enumerate(songs):

    print(f"{i}/{len(songs)}")

    # Convert to query parameters
    query = f"{song_title_artist} songtekst lyrics"

    base_url = "https://www.google.com/search?"


    options = webdriver.ChromeOptions()
    # options.add_argument('--headless')

    # Create a new Chrome WebDriver instance
    driver = webdriver.Chrome(options=options)
    # driver.maximize_window()

    # Navigate to a website
    driver.get(base_url)

    buttons = driver.find_elements(By.TAG_NAME, "button")

    # Iterate through the list and print the text of each button
    for index, button in enumerate(buttons, start=1):
        # print(f"Button {index}:")
        # print(f"Text: {button.text}")
        # print("-" * 30)

        if button.text == "Alles accepteren":
            button.click()



    # Locate all textarea elements on the page
    textareas = driver.find_elements(By.TAG_NAME, "textarea")
    query_textarea = None

    # Iterate through the list and print information about each textarea
    for index, textarea in enumerate(textareas, start=1):
        # print(f"Textarea {index}:")
        # print(f"Text Content: {textarea.get_attribute('value')}")
        # print(f"Title: {textarea.get_attribute('title')}")
        # print(f"Placeholder: {textarea.get_attribute('placeholder')}")
        # print("-" * 30)

        if textarea.get_attribute('title') == "Zoeken":
            # query_textarea = WebDriverWait(driver, 10).until(EC.element_to_be_selected(textarea))
            query_textarea = textarea





    query_textarea.send_keys(query)

    # Submit the form
    query_textarea.send_keys(Keys.RETURN)  # If pressing Enter submits the form

    # driver.implicitly_wait(5)

    try:
        # Wait for a specific element on the result page to be visible (adjust the selector as needed)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
    except Exception as e:
        print(f"Error waiting for the page: {e}")

    # Print the HTML of the result page
    # print(driver.page_source)


    # Print the page source
    # print(">>",driver.page_source)

    # Get the html page source of the current page
    page_source = driver.page_source

    # exit()

    soup = BeautifulSoup(page_source, 'html.parser')

    songtext_paragraphs = soup.find_all('body') # class_='songtext'
    # songtext_paragraphs = soup.find_all('p', class_='songtext') # class_='songtext'

    # print(songtext_paragraphs[0])
    # exit()

    song = song_title_artist
    print(song)

    if page_source.count('Bron:') < 1:
        print(f"{song} NOT FOUND")

    for p in songtext_paragraphs:
        text = p.text.replace("\n", " ").replace("  ", " ").upper()

        processed_text = re.sub(r'[^AEIOU\s]', '', text)

        # alternative, also remove whitespace
        processed_text = re.sub(r'[^AEIOU]', '', text)

        # Regex pattern to find 4 words with specific vowel sequences
        patterns = [
            r'\bAIEO\b\s+\bAIEO\b\s+\bAIEO\b\s+\bIAO\b',
            r'\bA\b\s+\bA\b\s+\bA\b\s+\bI\b\s+\bEE\b\s+\bA\b\s+\bA\b\s+\bA\b\s+\bI\b\s+\bEE\b\s+\bA\b\s+\bA\b\s+\bU\b',
            r'\bOEE\b\s+\bUI\b\s+\bO\b\s+\bE\b\s+\bAE\b\s+\bA\b\s+\bOI\b\s+\bEE\b\s+\bAE\b',
            r'\bO\b\s+\bI\b\s+\bI\b\s+\bO\b\s+\bI\b\s+\bI\b\s+\bOU\b\s+\bEE\b\s+\bEE\b'
        ]

        # alternative pattern without whitespace
        patterns = [
            # r'AIEOAIEOAIEOIAO',
            # r'AAAIEEAAAIEEAAU',
            # r'OEEUIOEAEAOIEEAE',
            # r'OIIOIIOUEEEE',
            r'OOIIOAAIEI'
        ]

        # Perform the search
        for i, pattern in enumerate(patterns):
            # print(processed_text)
            matches = re.findall(pattern, processed_text)

            if len(matches) > 0:
                print(text)
                print(f"{song} Found pattern {i + 1}")

    driver.quit()
    time.sleep(10)
    # break

    # Close the browser window
